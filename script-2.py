import openpyxl
import warnings
import requests
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from collections import defaultdict
from openpyxl.utils import column_index_from_string, get_column_letter
import re
import csv
# === CONFIGURATION ===
INPUT_FILE = "Blank MPS Batch Order Workbook.xlsx"
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# === MAIN ===
def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws1 = wb["Batch Metadata"]
    ws2 = wb["Audio Digitization Inventory"]
    setup_batch_metadata(ws1, ws2)
    extract_and_write_audio_metadata(ws1, ws2)
    OUTPUT_FILE = fill_batch_name(ws1,ws2)+".xlsx"
    
    
    wb.save(OUTPUT_FILE)
    print(f"\n✅ All changes saved to '{OUTPUT_FILE}'.")

# === SUBROUTINES ===
def setup_batch_metadata(ws1, ws2):
    ws1["C5"] = date.today().strftime("%m/%d/%Y")

    radio_button_prompt("What is the name of the collection?", {
        "1": {"Collection": "AWM Collection"},
        "2": {"Collection": "RECCO Collection"}
    }, {"B3": "Collection"}, ws1)

    radio_button_prompt("Who is the curator granting copy permission?", {
        "1": {"name": "Elizabeth Batiuk"},
        "2": {"name": "Peter Laurence"}
    }, {"B4": "name", "B6": "name"}, ws1)

    radio_button_prompt("What is the order type for this batch?", {
        "1": {"type": "Patron"},
        "2": {"type": "Standard"}
    }, {"B8": "type"}, ws1)

    print("Is this a rush order?\n1. Yes\n2. No")
    choice = input("Enter your choice: ").strip()
    while choice not in {"1", "2"}:
        choice = input("Invalid choice. Try again: ").strip()
    if choice == "1":
        date_prompt("What is the deadline for delivering a rush order", "C8", ws1)

    radio_button_prompt("What is the access level for this batch?", {
        "1": {"flag": "R"},
        "2": {"flag": "P"},
        "3": {"flag": "N"}
    }, {"B10": "flag"}, ws1)

    ws2["I3"] = ws1["B10"].value


def extract_and_write_audio_metadata(ws1, ws2):
    call_number_terms = prompt_for_call_numbers()

    # Step 2: Fetch all matching MODS records
    matched_mods_records = fetch_matching_mods_records(call_number_terms)

    if not matched_mods_records:
        print("❌ No matching records found for the given call numbers.")
        return
    
    # Namespace used throughout
    ns = {"mods": "http://www.loc.gov/mods/v3"}
    ROLE_TO_COLUMN = {
        "Actor": "N", "Adapter": "O", "Arranger": "P", "Author": "Q", "Commentator": "R",
        "Composer": "S", "Creator": "T", "Director": "V", "Instrumentalist": "X", "Interviewee": "Y",
        "Interviewer": "Z", "Librettist": "AA", "Lyricist": "AB", "Moderator": "AC",
        "Musical Director": "AD", "Musician": "AE", "Narrator": "AF", "Performer": "AG",
        "Publisher": "AH", "Rights Statement": "AI", "Singer": "AJ", "Speaker": "AK",
        "Storyteller": "AL", "Subject": "AM", "Vocalist": "AN"
    }
    
    def get_text(mods, xpath, ns):
        el = mods.find(xpath, ns)
        return el.text.strip() if el is not None and el.text else ""

    def get_text_from_element(el):
        return el.text.strip() if el is not None and el.text else ""
    
    def get_clean_finding_aid(mods, ns):
        record_info = mods.find(".//mods:recordInfo", ns)
        if record_info is not None:
            for rid in record_info.findall("mods:recordIdentifier", ns):
                if len(rid.attrib) == 0 and rid.text:
                    raw_id = rid.text.strip()
                    return raw_id.split('c')[0] if 'c' in raw_id else raw_id
        related_items = mods.findall(".//mods:relatedItem[@otherType='Finding Aid']", ns)
        for related in related_items:
            url_el = related.find(".//mods:url", ns)
            if url_el is not None and url_el.text:
                match = re.search(r"(mus\d+)", url_el.text.lower())
                if match:
                    return match.group(1)

        return ""
    def classify_format(call_number):
        call_number = call_number.upper()
        if "SC" in call_number: return "COMPACT CASSETTE"
        if "RL" in call_number: return '1/4" OPEN REEL AUDIO'
        if "DAT" in call_number: return "DAT"
        if "CD" in call_number: return "COMPACT DISC"
        if "ER" in call_number: return "DIGITAL AUDIO FILE"
        if "LP" in call_number or re.search(r"AWM\s*45[-\s]?\d{4}", call_number): return "LP"
        return ""
    def clean_item_label(item_label):
        match = re.search(r"(AWM\s+[A-Z]+\s+\d+)", item_label.upper())
        if match:
            return match.group(1)
        return ""

    radio_button_prompt("Should this batch be available for download?", {
        "1": {"flag": "TRUE"},
        "2": {"flag": "FALSE"}
    }, {"G3": "flag"}, ws2)

    if ws2["G3"].value == "FALSE":
        radio_button_prompt("Why is this batch not available for download?", {
            "1": {"option": "Harvard Policy"}, "2": {"option": "License"},
            "3": {"option": "Risk Assessment"}, "4": {"option": "Statute"},
            "5": {"option": "Copyright"}
        }, {"H3": "option"}, ws2)

    radio_button_prompt("What is the archival file storage class for this item?", {
        "1": {"option": "Sensitive"}, "2": {"option": "Basic"},
        "3": {"option": "Large"}, "4": {"option": "Archival"}
    }, {"J3": "option"}, ws2)
    
    download_flag = ws2["G3"].value
    download_reason = ws2["H3"].value if download_flag == "FALSE" else ""
    access_flag = ws2["I3"].value
    storage_class = ws2["J3"].value
    due_date = (datetime.today() + timedelta(weeks=2)).strftime("%m/%d/%Y")

    # === Loop over each matching mods record ===
    row_index = 3
    has_table_of_content = {}
    for mods in matched_mods_records:
        item_label = get_text(mods, ".//mods:shelfLocator", ns) or get_container_location(mods, ns)
        item_label = clean_item_label(item_label)
        if not item_label:
            continue
        has_table_of_content[item_label] = extract_table_of_contents(mods, ns)
        title = get_text(mods, "mods:titleInfo/mods:title", ns)
        subtitle = get_text(mods, "mods:titleInfo/mods:subTitle", ns)
        part_number = get_text(mods, "mods:titleInfo/mods:partNumber", ns)
        finding_aid = get_clean_finding_aid(mods, ns)
        full_title = f"{title}: {subtitle}" if subtitle else title
        full_title = f"{full_title}: {part_number}" if part_number else full_title

        date_issued = get_text(mods, "mods:originInfo/mods:dateIssued", ns)
        publisher = get_text(mods, "mods:originInfo/mods:publisher", ns)
        hollis_number = get_text(mods, "mods:recordInfo/mods:recordIdentifier[@source='MH:ALMA']", ns)
        genres = [g.text.strip() for g in mods.findall("mods:genre", ns) if g.text]
        format_value = classify_format(item_label)

        def extract_creators():
            result = defaultdict(list)
            for name in mods.findall("mods:name[@type='personal']", ns):
                parts = [part.text.strip() for part in name.findall("mods:namePart", ns) if part.text]
                full_name = ", ".join(parts)
                roles = [r.text.strip().rstrip(".").title() for r in name.findall("mods:role/mods:roleTerm[@type='text']", ns) if r.text]
                for role in roles or [None]:
                    col = ROLE_TO_COLUMN.get(role)
                    if col:
                        result[col].append(full_name)
                    else:
                        result["T"].append(f"{full_name} – {role}" if role else full_name)
            for corp in mods.findall("mods:name[@type='corporate']", ns):
                corp_name = get_text_from_element(corp.find("mods:namePart", ns))
                if corp_name:
                    result["T"].append(corp_name)
            return result

        creator_columns = extract_creators()

        ws2.cell(row=row_index, column=1, value=item_label)
        ws2.cell(row=row_index, column=3, value=item_label)
        ws2.cell(row=row_index, column=4, value=format_value)
        ws2.cell(row=row_index, column=5, value=hollis_number)
        ws2.cell(row=row_index, column=6, value=finding_aid)
        ws2.cell(row=row_index, column=7, value=download_flag)
        ws2.cell(row=row_index, column=8, value=download_reason)
        ws2.cell(row=row_index, column=9, value=access_flag)
        ws2.cell(row=row_index, column=10, value=storage_class)
        ws2.cell(row=row_index, column=12, value=full_title)
        for col_letter, names in creator_columns.items():
            col_num = column_index_from_string(col_letter)
            ws2.cell(row=row_index, column=col_num, value="; ".join(names))
        ws2.cell(row=row_index, column=21, value=date_issued)
        ws2.cell(row=row_index, column=23, value="; ".join(genres) if genres else "")
        ws2.cell(row=row_index, column=34, value=publisher)
        ws2.cell(row=row_index, column=41, value=full_title)
        ws2.cell(row=row_index, column=42, value=due_date)
        row_index+=1

    items_with_toc = [(label, toc) for label, toc in has_table_of_content.items() if toc]
    if items_with_toc:
        print("The following items have a table of contents:")
        for label, toc in items_with_toc:
            print(f"{label}: {toc[:100]}{'...' if len(toc) > 100 else ''}")  # preview
    print("Would you like to view the table of contents for these items in order to populate the Audio Playlist Track Data workbook?\n1. Yes\n2. No")
    choice = input("Enter your choice: ").strip()
    while choice not in {"1", "2"}:
        choice = input("Invalid choice. Try again: ").strip()
    if choice == "1":
        with open("table_of_contents_export.csv", "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Item Label", "Table of Contents"])

            for label, toc in has_table_of_content.items():
                if toc:  # only include entries with TOC text
                    writer.writerow([label, toc])
            print("Table of contents saved to table_of_contents_export.csv")
            

def radio_button_prompt(question, options, fill_map, ws):
    print(question)
    for key, data in options.items():
        print(f"{key}. {next(iter(data.values()))}")
    choice = input("Enter your choice: ").strip()
    while choice not in options:
        choice = input("Invalid choice. Try again: ").strip()
    selected = options[choice]
    for cell, field_key in fill_map.items():
        ws[cell] = selected[field_key]

def date_prompt(question, cell, ws):

    print(question + " (format: MM/DD/YYYY)")
    while True:
        user_input = input("Enter the date: ").strip()
        try:
            parsed_date = datetime.strptime(user_input, "%m/%d/%Y")
            ws[cell] = parsed_date.strftime("%m/%d/%Y")
            break
        except ValueError:
            print("❌ Invalid format. Please enter date as MM/DD/YY.")

def is_valid_call_number(call):
    call = call.strip().upper()
    patterns = [
        r"^AWM [A-Z]+ \d{1,5}$",            # Single call number
        r"^AWM [A-Z]+ \d{1,5}-\d{1,5}$",    # Range
        r"^AWM SPEC COLL \d{1,5}$",         # Collection
        r"^AWM SPEC COLL \d{1,5}-\d{1,5}$"  # Collection Range
    ]
    return any(re.match(p, call) for p in patterns)

def expand_call_range(call):
    """
    Expands AWM LP 1900-1902 → ["AWM LP 1900", "AWM LP 1901", "AWM LP 1902"]
    """
    prefix_match = re.match(r"^(AWM(?: [A-Z]+)*?) (\d+)-(\d+)$", call)
    if not prefix_match:
        return [call]
    base, start, end = prefix_match.groups()
    return [f"{base} {i}" for i in range(int(start), int(end)+1)]

def prompt_for_call_numbers():
    while True:
        raw_input = input("Enter call number(s):\n"
                          "- Single (e.g. AWM LP 2030)\n"
                          "- Comma-separated (e.g. AWM LP 2030, AWM SC 1010)\n"
                          "- Range (e.g. AWM LP 2000–2010)\n"
                          "- Collection (e.g. AWM SPEC COLL 98)\n> ").strip()

        raw_input = raw_input.replace("–", "-").upper()  # Normalize en dash and uppercase
        entries = [x.strip() for x in raw_input.split(",")]

        if all(is_valid_call_number(cn) for cn in entries):
            expanded = []
            for entry in entries:
                expanded.extend(expand_call_range(entry))
            return expanded
        else:
            print("❌ One or more call numbers are invalid. Please try again.\n")

def fetch_matching_mods_records(call_number_terms):
  
    if not call_number_terms:
        return []

    
    
    ns = {
        "mods": "http://www.loc.gov/mods/v3",
        "librarycloud": "http://hul.harvard.edu/ois/xml/ns/librarycloud"
    }
    matched_mods = []
    for term in call_number_terms:
        api_url = f"https://api.lib.harvard.edu/v2/items?q={term}&limit=20000"
        response = requests.get(api_url)
        if response.status_code != 200:
            print(f"❌ Error fetching data for term: {term}")
            continue
        root = ET.fromstring(response.content)
        matched_mods+=filter_matched_mods(term, root, ns)
    return matched_mods

def filter_matched_mods(term, root, ns):
    matched_mods_term = []
    for mods in root.findall(".//mods:mods", ns):
        shelf_text = get_container_location(mods, ns)
        shelf_locator_el = mods.find(".//mods:shelfLocator", ns)
        if shelf_locator_el is not None and shelf_locator_el.text:
            shelf_text += shelf_locator_el.text.upper().strip()            
        
        unit_id_elements = mods.findall(".//mods:identifier", ns)
        unit_text = " ".join(el.text.strip().upper() for el in unit_id_elements if el is not None and el.text)
        parts = term.upper().split()
        if all(part in shelf_text for part in parts) or all(part in unit_text for part in parts):
            matched_mods_term.append(mods)
    return matched_mods_term
    

def get_container_location(mods, ns):
    call_number=""
    for location in mods.findall("mods:location", ns):
            for phys in location.findall("mods:physicalLocation", ns):
                if phys.attrib.get("type") == "container":
                    call_number += phys.text.strip() if phys.text else None
    
    return call_number                   

def fill_batch_name(ws1, ws2):
    def batch_name_call_number(ws2):
            entries = []
            for row in range(3, ws2.max_row + 1):
                val = ws2[f"C{row}"].value
                if not val:
                    continue  # Skip blank rows

                parts = val.strip().split()
                if len(parts) != 3:
                    continue  # Skip malformed entries

                try:
                    number = int(parts[2])
                except ValueError:
                    continue  # Skip if third part isn't a number (e.g. "and")
                entries.append(parts)

            # === Process valid entries ===
            if not entries:
                return ""
            if len(entries) == 1:
                return "_".join(entries[0])+"_"  # e.g., AWM_LP_2078

            sorted_entries = sorted(entries, key=lambda x: int(x[2]))
            awm = sorted_entries[0][0]
            type1, num1 = sorted_entries[0][1], sorted_entries[0][2]
            type2, num2 = sorted_entries[-1][1], sorted_entries[-1][2]

            if type1 == type2:
                return f"{awm}_{type1}_{num1}_{num2}_"
            else:
                return f"{awm}_{type1}_{num1}_{type2}_{num2}_"
    
    prefix = "hcl-aps_batch_dropoff_"
    if ws1["B8"].value == "Patron":
        prefix += "Requests_"
    formatted_call = batch_name_call_number(ws2)
    date = datetime.strptime(ws1["C5"].value, "%m/%d/%Y").strftime("%Y%m%d")
    initials = "EB" if ws1["B3"].value == "AWM Collection" else "PL"
    batch_name = prefix + formatted_call + date + initials
    ws1["B11"] = batch_name
    return batch_name

def extract_table_of_contents(mods, ns):
    toc_elements = mods.findall(".//mods:tableOfContents", ns)
    toc_texts = [el.text.strip() for el in toc_elements if el.text]
    return "\n".join(toc_texts) if toc_texts else None
if __name__ == "__main__":
    main()
